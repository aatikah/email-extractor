import openpyxl

def extract_email_column(file_path, sheet_name='Form Responses 1', email_column_name='Email Address'):
    # Load the workbook and select the sheet
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    
    # Find the email column index by matching the header
    email_column_index = None
    for cell in sheet[1]:  # Assuming the first row contains headers
        if cell.value == email_column_name:
            email_column_index = cell.column
            break
    
    if email_column_index is None:
        raise ValueError(f"Column '{email_column_name}' not found in the sheet.")
    
    # Extract email data
    emails = []
    for row in sheet.iter_rows(min_row=2, max_col=email_column_index, max_row=sheet.max_row):
        email_cell = row[email_column_index - 1]
        if email_cell.value:
            emails.append(email_cell.value)
    
    return emails

# Example usage:
file_path = 'em.xlsx'  # Path to your Excel file
output_file = 'emails.txt'  # Output file name

emails = extract_email_column(file_path)

# Print each email on a new line
for email in emails:
    print(email)
    
# Write emails to the text file, each on a new line
with open(output_file, 'w') as f:
    for email in emails:
        f.write(email + '\n')

print(f"Emails have been written to {output_file}")
